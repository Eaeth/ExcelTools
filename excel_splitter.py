# -*- coding: utf-8 -*-

import openpyxl
import os
import copy
from openpyxl.utils import get_column_letter
from zipfile import BadZipFile
from util import get_date_str
import win32com.client
import pywintypes


class ExcelSplitter:
    def __init__(self, input_file, start_row, end_row, key_column, out_put_path, password):
        self.input_file = input_file
        self.start_row = start_row
        self.end_row = end_row
        self.key_column = key_column
        self.out_put_path = out_put_path
        self.password = password
        self.data_dict = {}

    def split_excel_by_column(self):

        # 获取工作簿和工作表对象
        try:
            wb_in = openpyxl.load_workbook(self.input_file)
        except BadZipFile:
            return f"文件 '{self.input_file}' 带有密码，请去除密码后再处理!"

        if self.password != None:
            try:
                excel = win32com.client.gencache.EnsureDispatch(
                    'Excel.Application')
            except pywintypes.com_error:
                try:
                    excel = win32com.client.gencache.EnsureDispatch(
                        'KWPS.Application')
                    excel.Quit()
                    return "请打开 WPS Office 兼容选项\n位置: 设置->组件管理->文件格式关联...->兼容第三方系统和软件"
                except pywintypes.com_error:
                    return "未找到 Microsoft Excel 或 WPS Office,请先安装！"

        ws_in = wb_in.active
        ws_in_max_column = ws_in.max_column
        ws_in_max_row = ws_in.max_row
        if self.end_row > ws_in_max_row:
            self.end_row = ws_in_max_row

        # 获取开始到结束行的值
        col_values = [ws_in.cell(row=i, column=self.key_column).value for i in range(
            self.start_row, self.end_row+1)]
        # 创建一个字典，键为第key_column列的值，值为包含该值的所有行的列表
        for i, value in enumerate(col_values):
            if value in self.data_dict:
                self.data_dict[value].append(i + self.start_row)
            else:
                self.data_dict[value] = [i + self.start_row]
        # 对于每个键值对，创建一个新的工作簿，将包含该值的行拷贝到新工作簿
        for key, rows in self.data_dict.items():
            if key == None:
                continue

            wb_out = openpyxl.Workbook()
            ws_out = wb_out.active
            ws_out.title = ws_in.title

            merged_ranges = set()

            # Copy first two rows
            if self.start_row != 1:
                for row in range(1, self.start_row):
                    ws_out.row_dimensions[row].height = ws_in.row_dimensions[row].height
                    for col in range(1, ws_in_max_column+1):
                        ws_out.column_dimensions[get_column_letter(
                            col)].width = ws_in.column_dimensions[get_column_letter(col)].width

                        # 设置单元格格式
                        source_cell = ws_in.cell(row, col)
                        target_cell = ws_out.cell(row, col)
                        self.copy_cell(source_cell, target_cell)
                        # 如果当前单元格是合并单元格，则在 ws_out 中也合并相应的单元格
                        if source_cell.coordinate in ws_in.merged_cells:
                            self.merge_cell(ws_in, source_cell,
                                            col, row, merged_ranges)

            # 拷贝包含该值的行
            for i, row in enumerate(rows):
                ws_out.row_dimensions[i +
                                      self.start_row].height = ws_in.row_dimensions[row].height
                for col in range(1, ws_in_max_column+1):
                    ws_out.column_dimensions[get_column_letter(
                        col)].width = ws_in.column_dimensions[get_column_letter(col)].width

                    # 设置单元格格式
                    source_cell = ws_in.cell(row=row, column=col)
                    target_cell = ws_out.cell(row=i+self.start_row, column=col)
                    self.copy_cell(source_cell, target_cell)
                    # 如果当前单元格是合并单元格，则在 ws_out 中也合并相应的单元格
                    if source_cell.coordinate in ws_in.merged_cells:
                        self.merge_cell(ws_in, source_cell, col,
                                        i+self.start_row, merged_ranges)

            # 跳过中间空白
            del_num_rows = self.end_row - self.start_row - len(rows) + 1

            if ws_in_max_row > self.end_row:
                # 拷贝剩余内容
                for row_in in range(self.end_row+1, ws_in_max_row+1):
                    row_out = row_in - del_num_rows
                    ws_out.row_dimensions[row_out].height = ws_in.row_dimensions[row_in].height
                    for col in range(1, ws_in_max_column+1):
                        ws_out.column_dimensions[get_column_letter(
                            col)].width = ws_in.column_dimensions[get_column_letter(col)].width
                        # 设置单元格格式
                        source_cell = ws_in.cell(row_in, col)
                        target_cell = ws_out.cell(row_out, col)
                        self.copy_cell(source_cell, target_cell)
                        # 如果当前单元格是合并单元格，则在 ws_out 中也合并相应的单元格
                        if source_cell.coordinate in ws_in.merged_cells:
                            self.merge_cell(ws_in, source_cell,
                                            col, row_out, merged_ranges)

            # 处理合并的单元格
            for cell_range in merged_ranges:
                ws_out.merge_cells(cell_range)

            # 保存拷贝后的文档
            file_name = os.path.splitext(os.path.basename(self.input_file))[0]
            date = get_date_str()
            save_path = f"{self.out_put_path}/{key}{file_name}{date}.xlsx"
            try:
                wb_out.save(save_path)
            except PermissionError:
                wb_in.close()
                if self.password != None:
                    excel.Quit()
                return f"保存文件 {save_path} 失败，请检查是否被其他程序占用"
            wb_out.close()
            if self.password != None:
                workbook = excel.Workbooks.Open(save_path)
                workbook.Password = self.password
                # 保存Excel文件
                workbook.Save()
                # 关闭Excel文件和应用程序
                workbook.Close()

        wb_in.close()
        if self.password != None:
            excel.Quit()

    def copy_cell(self, cell_in, cell_out):
        # 复制单元格的值、样式、边框、对齐方式、字体、填充颜色
        cell_out.value = cell_in.value
        if cell_in.has_style:
            cell_out._style = copy.copy(cell_in._style)
            cell_out.font = copy.copy(cell_in.font)
            cell_out.border = copy.copy(cell_in.border)
            cell_out.fill = copy.copy(cell_in.fill)
            cell_out.number_format = copy.copy(cell_in.number_format)
            cell_out.protection = copy.copy(cell_in.protection)
            cell_out.alignment = copy.copy(cell_in.alignment)

    def merge_cell(self, ws_in, cell_in, col_out, row_out, merged_ranges):
        # 目前都是按行拷贝，列不变
        # 获取源单元格所属合并单元格的范围
        for range_string in ws_in.merged_cells.ranges:
            min_row, max_row, min_col, max_col = range_string.min_row, range_string.max_row, range_string.min_col, range_string.max_col

            # 如果源单元格在该合并单元格内，则在输出工作表中也合并相应的单元格
            if min_row <= cell_in.row <= max_row and min_col <= cell_in.column <= max_col:
                tl_coord = f"{get_column_letter(min_col)}{row_out+min_row-cell_in.row}"
                br_coord = f"{get_column_letter(max_col)}{row_out+max_row-cell_in.row}"
                cell_range = f"{tl_coord}:{br_coord}"

                # 如果该合并单元格已经被合并，则跳过
                if cell_range in merged_ranges:
                    continue
                merged_ranges.add(cell_range)
                return
