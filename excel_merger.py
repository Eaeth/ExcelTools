# -*- coding: utf-8 -*-

import openpyxl
import os
import copy
from openpyxl.utils import get_column_letter
from zipfile import BadZipFile


class ExcelMerger:
    def __init__(self, input_path, head_rows, key_column, out_put_path):
        self.input_path = input_path
        self.head_rows = head_rows
        self.key_column = key_column
        self.out_put_path = out_put_path
        self.merged_ranges = set()

    def merge_excels(self):
        # 创建wb_out
        wb_out = openpyxl.Workbook()

        ws_out = wb_out.active
        head_row = False
        head_row_file = ''

        # 读取excel文件列表
        for filename in os.listdir(self.input_path):
            if not (filename.endswith('.xls') or filename.endswith('.xlsx')):
                continue

            # 打开文件
            filepath = os.path.join(self.input_path, filename)
            try:
                wb_in = openpyxl.load_workbook(filepath)
            except BadZipFile:
                return f"文件 '{filename}' 带有密码，请去除密码后再处理!"
            except PermissionError:
                return f"读取文件 {filename} 失败，请检查是否被其他程序占用"
            ws_in = wb_in.active

            # 如果head_row标记为False，则拷贝表头
            if not head_row:
                for row in range(1, self.head_rows+1):
                    ws_out.row_dimensions[row].height = ws_in.row_dimensions[row].height
                    for col in range(1, ws_in.max_column+1):
                        cell_in = ws_in.cell(row=row, column=col)
                        cell_out = ws_out.cell(row=row, column=col)
                        self.copy_cell(cell_in, cell_out)
                        # 如果当前单元格是合并单元格，则在 ws_out 中也合并相应的单元格
                        if cell_in.coordinate in ws_in.merged_cells:
                            self.merge_cell(ws_in, cell_in, col, row)
                        # 拷贝列宽
                        ws_out.column_dimensions[get_column_letter(
                            col)].width = ws_in.column_dimensions[get_column_letter(col)].width
                head_row = True
                ws_out.title = ws_in.title
                head_row_file = filepath
            else:
                # 检查表头是否一致
                for row in range(1, self.head_rows+1):
                    for col in range(1, ws_in.max_column+1):
                        cell_in = ws_in.cell(row=row, column=col)
                        cell_out = ws_out.cell(row=row, column=col)
                        if cell_in.value != cell_out.value:
                            return ("表头不一致，位置：{0}{1} \n{2} 文件是 '{3}'\n{4} 文件是 '{5}'"
                                    .format(get_column_letter(col), row, os.path.normpath(filepath),
                                            cell_in.value, os.path.normpath(head_row_file), cell_out.value))

            # 从 head_rows+1 行开始遍历数据，直到遇到空单元格为止
            for row in range(self.head_rows+1, ws_in.max_row+1):
                cell_in = ws_in.cell(row=row, column=self.key_column)
                if cell_in.value is None:
                    break

                # 在 ws_out 中找到下一个空行，并将当前行数据复制到该行
                row_out = ws_out.max_row + 1
                ws_out.row_dimensions[row_out].height = ws_in.row_dimensions[row].height
                for col in range(1, ws_in.max_column+1):
                    cell_in = ws_in.cell(row=row, column=col)
                    cell_out = ws_out.cell(row=row_out, column=col)
                    self.copy_cell(cell_in, cell_out)

                    # 如果当前单元格是合并单元格，则在 ws_out 中也合并相应的单元格
                    if cell_in.coordinate in ws_in.merged_cells:
                        self.merge_cell(ws_in, cell_in, col, row_out)
                    # 拷贝列宽
                    ws_out.column_dimensions[get_column_letter(
                        col)].width = ws_in.column_dimensions[get_column_letter(col)].width
            # 关闭文件
            wb_in.close()

        # 处理合并的单元格
        for cell_range in self.merged_ranges:
            ws_out.merge_cells(cell_range)

        # 保存输出文件
        try:
            wb_out.save(self.out_put_path)
        except PermissionError:
            return f"保存文件 {self.out_put_path} 失败，请检查是否被其他程序占用"
        wb_out.close()

    def merge_cell(self, ws_in, cell_in, col_out, row_out):
        # 目前都是按行拷贝，列不变
        # 获取源单元格所属合并单元格的范围
        for range_string in ws_in.merged_cells.ranges:
            min_row, max_row, min_col, max_col = range_string.min_row, range_string.max_row, range_string.min_col, range_string.max_col

            # 如果源单元格在该合并单元格内，则在输出工作表中也合并相应的单元格
            if min_row <= cell_in.row <= max_row and min_col <= cell_in.column <= max_col:
                tl_coord = f"{get_column_letter(min_col)}{row_out+min_row-cell_in.row}"
                br_coord = f"{get_column_letter(max_col)}{row_out+max_row-cell_in.row}"
                cell_range = f"{tl_coord}:{br_coord}"

                # 如果该合并单元格已经被合并，则跳过
                if cell_range in self.merged_ranges:
                    continue
                self.merged_ranges.add(cell_range)
                return

    def copy_cell(self, cell_in, cell_out):
        # 复制单元格的值、样式、边框、对齐方式、字体、填充颜色
        cell_out.value = cell_in.value
        if cell_in.has_style:
            cell_out._style = copy.copy(cell_in._style)
            cell_out.font = copy.copy(cell_in.font)
            cell_out.border = copy.copy(cell_in.border)
            cell_out.fill = copy.copy(cell_in.fill)
            cell_out.number_format = copy.copy(cell_in.number_format)
            cell_out.protection = copy.copy(cell_in.protection)
            cell_out.alignment = copy.copy(cell_in.alignment)
